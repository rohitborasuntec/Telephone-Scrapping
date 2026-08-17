"""
Neighbour Detection Pipeline (v6)
==================================
Combines OS Places API (address/UPRN lookup) with OS NGD API (building AND
land parcel polygons) to identify TRUE boundary-touching neighbours for a
target property, then applies the client's business rules.

Layered defences against false-positive "neighbours":
  1. Buffered .intersects() as a coarse first pass (buildings + land parcels).
  2. shares_real_boundary(): the intersection must be an actual LINE of
     meaningful length, not just a single touching corner POINT.
  3. is_clean_two_party_boundary(): rejects a shared boundary if a THIRD
     property's parcel sits right at the MIDPOINT of that boundary — a
     signature of a multi-way "pinch point" where 3+ plots meet at a
     corner. Confirmed threshold: genuine neighbours had nearest-third-
     party distances of ~3.9-9m+; pinch-point false positives were ~1.5m.
     PINCH_POINT_CHECK_DISTANCE_M is set at 2.0m to sit in that gap.
  4. sanity_check_distance(): final straight-line distance check between
     the target's and candidate's actual address points. Catches cases
     the polygon logic can't — e.g. cul-de-sac ("Close") developments
     where two unrelated properties both border the same large shared/
     communal land parcel (a turning circle, verge, amenity strip), which
     makes them register as "touching" even though they are nowhere near
     each other in reality.

STATUS / STILL TUNING:
  MIN_THIRD_PARTY_AREA_SQM and MAX_PLAUSIBLE_NEIGHBOUR_DISTANCE_M are
  starting values based on the test addresses worked through so far, not
  exhaustively validated. Recommended before going live: build a table of
  15-20 known addresses (using client-style aerial screenshots to confirm
  expected neighbours) and run find_neighbours() on all of them after any
  threshold change, checking nothing that used to pass now fails and vice
  versa. Tune one constant at a time.

WHAT'S NEW IN v6:
  - Proper `logging` module setup (console + rotating-by-run log file)
    replaces the old bare `print()` calls throughout the pipeline.
  - Two CSV outputs, written incrementally (row-by-row / address-by-address)
    so a crash partway through a big batch doesn't lose everything already
    processed:
      * NEIGHBOURS_CSV_PATH  -> one row per confirmed neighbour only.
      * FULL_RESULTS_CSV_PATH -> everything the pipeline produced for every
        candidate it looked at (neighbours AND rejected candidates AND a
        row for addresses that returned zero candidates), so you can audit
        *why* something was or wasn't counted as a neighbour.

WHAT'S NEW IN v7:
  - match_address_to_polygon() no longer fails silently: on a miss it now
    logs the nearest address point distance it *did* find (or that none
    existed at all within the search radius), so "no address point found
    inside polygon" rejections are actually diagnosable instead of a dead
    end.
  - Same-road filter: a touching candidate whose THOROUGHFARE_NAME differs
    from the target's is no longer accepted as a neighbour even if it
    geometrically touches and passes the distance sanity check. This
    catches the case where a target's REAR garden parcel touches a
    property on a completely different street (a back-to-back garden
    relationship, not a true neighbour) — the pinch-point/distance checks
    alone don't reliably catch this because the touching parcel is real
    and often not that far away. These are still recorded in the full
    results CSV as "rejected", with the reason spelled out, so nothing is
    silently dropped.
  - unmatched_addresses.csv: any input address that fails to geocode (no
    results, or best match below the 0.7 threshold) is now captured with
    whatever partial match info OS Places did return, so failed inputs are
    easy to triage instead of showing up as a blank row.

WHAT'S NEW IN v8:
  - Target and neighbour coordinates are now converted from the OS grid
    (EPSG:27700 / British National Grid) to WGS84 lat/lon via `pyproj`,
    the same approach as the reference snippet — reusing one cached
    Transformer per CRS rather than rebuilding it on every address point.
  - Alongside the incremental CSVs, the run now also writes two formatted
    .xlsx workbooks (openpyxl) once the batch finishes: Title Case column
    headers with a space between words, bold header row, frozen header,
    and auto-sized columns. The CSVs remain the row-by-row audit trail
    written during the run (so a crash mid-batch still leaves usable
    output); the .xlsx files are the polished deliverable built from them
    at the end.

Requirements:
    pip install requests shapely pandas python-dotenv pyproj openpyxl

You will need an OS Data Hub API key on the Premium Plan (free tier /
development mode is fine for testing) with these APIs added to the project:
    - OS Places API
    - OS NGD API - Features
"""

import csv
import logging
import os
import uuid
from datetime import datetime

import requests
from dotenv import load_dotenv
from pyproj import Transformer
from shapely.geometry import shape, Point

load_dotenv(override=True)
OS_API_KEY = os.getenv('API_KEY')

if not OS_API_KEY:
    raise ValueError("API_KEY not set in environment variables")

PLACES_BASE = "https://api.os.uk/search/places/v1"
NGD_BASE = "https://api.os.uk/features/ngd/ofa/v1/collections"

BUFFER_M = 1.0                        # slack distance for polygon-polygon adjacency test
SEARCH_PAD_M = 30.0                   # bbox padding around the target property, in metres
ADDRESS_POINT_SLACK_M = 2.0           # buffer applied to a polygon before checking address points
ADDRESS_MATCH_FALLBACK_M = 6.0        # nearest-address-point fallback distance
LAND_CONTAINMENT_FALLBACK_M = 5.0     # nearest-land-parcel fallback distance (target's own parcel)
MIN_SHARED_BOUNDARY_LENGTH_M = 0.3    # minimum length to count as a real shared edge (not a corner)
PINCH_POINT_CHECK_DISTANCE_M = 2.0    # confirmed against real data — see notes above
MIN_THIRD_PARTY_AREA_SQM = 30.0       # ignore tiny driveway/path slivers in the pinch-point check
MAX_PLAUSIBLE_NEIGHBOUR_DISTANCE_M = 30.0  # final address-point sanity check — tune against test set
MAX_NEIGHBOURS_BEFORE_ABORT = 4

CRS_TEMPLATE = "http://www.opengis.net/def/crs/EPSG/0/{}"

# Classification codes that should be EXCLUDED even though they start with 'R'
# (flats, maisonettes, multi-dwelling buildings). Confirm/expand this list
# against the OS classification scheme document before going live.
EXCLUDED_RESIDENTIAL_SUBTYPES = {
    "RD06",  # Flat (converted)
    "RD07",  # Flat (Purpose Built)
    "RD08",  # Flat (Maisonette) -- confirm exact codes with OS scheme doc
}

srs = "EPSG:27700"  # set dynamically from the geocode response

# ---------------------------------------------------------------------------
# OUTPUT / LOGGING PATHS
# ---------------------------------------------------------------------------
INPUT_FILE_PATH = r'c:\Users\hp\Downloads\final_results_new.csv'

OUTPUT_DIR = "Output_Neighbour"
LOG_DIR = "Logs_Neighbour"
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(LOG_DIR, exist_ok=True)

RUN_STAMP = datetime.now().strftime('%Y%m%d_%H%M%S')
NEIGHBOURS_CSV_PATH = os.path.join(OUTPUT_DIR, f"neighbours_{RUN_STAMP}.csv")
FULL_RESULTS_CSV_PATH = os.path.join(OUTPUT_DIR, f"full_results_{RUN_STAMP}.csv")
UNMATCHED_CSV_PATH = os.path.join(OUTPUT_DIR, f"unmatched_addresses_{RUN_STAMP}.csv")
LOG_FILE_PATH = os.path.join(LOG_DIR, f"pipeline_{RUN_STAMP}.log")

NEIGHBOURS_CSV_FIELDS = [
    "Reference Id",
    "Target Uprn", "Target Address", "Target Building Toid", "Target Land Toid",
    "Target Classification", "Target Latitude", "Target Longitude",
    "Neighbour Uprn", "Neighbour Address", "Neighbour Classification",
    "Neighbour Toid", "Neighbour Latitude", "Neighbour Longitude", "Distance M",
]

FULL_RESULTS_CSV_FIELDS = [
    "Reference Id",
    "Target Uprn", "Target Address", "Target Building Toid", "Target Land Toid",
    "Target Classification", "Target Latitude", "Target Longitude",
    "Aborted Multi Dwelling", "Neighbour Count",
    "Record Type",       # "Neighbour" | "Rejected" | "None Found" | "Error"
    "Uprn", "Address", "Classification", "Toid", "Latitude", "Longitude",
    "Distance M", "Reason",
]

UNMATCHED_CSV_FIELDS = [
    "Reference Id", "Input Address", "Raw Results Count", "Best Match Address", "Best Match Score",
]

# ---------------------------------------------------------------------------
# LOGGING SETUP
# ---------------------------------------------------------------------------
logger = logging.getLogger("neighbour_pipeline")
logger.setLevel(logging.DEBUG)
logger.propagate = False

if not logger.handlers:
    _fmt = logging.Formatter(
        "%(asctime)s | %(levelname)-7s | %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
    )

    _console_handler = logging.StreamHandler()
    _console_handler.setLevel(logging.INFO)   # keep console less noisy
    _console_handler.setFormatter(_fmt)

    _file_handler = logging.FileHandler(LOG_FILE_PATH, encoding="utf-8")
    _file_handler.setLevel(logging.DEBUG)     # full detail goes to file
    _file_handler.setFormatter(_fmt)

    logger.addHandler(_console_handler)
    logger.addHandler(_file_handler)


# ---------------------------------------------------------------------------
# STEP 1 — Geocode the target address
# ---------------------------------------------------------------------------
def geocode_address(address_text, diagnostics=None):
    """Geocode a free-text address via OS Places /find.

    `diagnostics`, if passed a dict, is populated with whatever partial
    match info is available even on failure (zero results, or a match
    below the 0.7 threshold) — used to build unmatched_addresses.csv so
    failed inputs are triageable instead of showing up as a blank row.
    """
    global srs
    url = f"{PLACES_BASE}/find"
    params = {"query": address_text, "maxresults": 1, "key": OS_API_KEY}

    logger.debug("Geocoding address: %s", address_text)
    r = requests.get(url, params=params)
    r.raise_for_status()
    data = r.json()
    srs = data.get("header", {}).get("output_srs", srs)
    results = data.get("results", [])

    if diagnostics is not None:
        diagnostics["raw_results_count"] = len(results)

    if not results:
        logger.warning("No geocode results for address: %s", address_text)
        return None

    try:
        best = results[0].get("DPA")
        is_match = best.get("MATCH")
        if diagnostics is not None:
            diagnostics["best_match_address"] = best.get("ADDRESS")
            diagnostics["best_match_score"] = is_match
        if is_match >= 0.7:
            return best
        logger.warning(
            "Geocode match score %.2f below threshold for: %s (best candidate: %s)",
            is_match, address_text, best.get("ADDRESS"),
        )
    except Exception:
        logger.exception("Unexpected geocode response shape for: %s", address_text)
        return None

    return None


def _crs():
    return CRS_TEMPLATE.format(srs.split(":")[-1])


# ---------------------------------------------------------------------------
# COORDINATE CONVERSION — OS grid (e.g. EPSG:27700 / British National Grid)
# to WGS84 lat/lon, same approach as the reference snippet. Transformers are
# cached per CRS code since building one is comparatively expensive and the
# CRS is the same for every point within a run.
# ---------------------------------------------------------------------------
_transformer_cache = {}


def get_transformer(crs_code):
    if crs_code not in _transformer_cache:
        _transformer_cache[crs_code] = Transformer.from_crs(crs_code, "EPSG:4326")
    return _transformer_cache[crs_code]


def xy_to_latlon(x, y, crs_code=None):
    """Convert an OS grid (X, Y) pair to (latitude, longitude) in WGS84.
    Defaults to the CRS most recently reported by the OS Places API
    (the module-level `srs`, e.g. 'EPSG:27700')."""
    crs_code = crs_code or srs
    transformer = get_transformer(crs_code)
    lat, lon = transformer.transform(x, y)
    return round(lat, 6), round(lon, 6)


# ---------------------------------------------------------------------------
# STEP 2 — Fetch polygons by TOID / by bbox, for a given NGD collection
# ---------------------------------------------------------------------------
def get_polygon_by_toid(collection, toid):
    url = f"{NGD_BASE}/{collection}/items"
    params = {
        "key": OS_API_KEY,
        "filter": f"toid='{toid}'",
        "filter-lang": "cql-text",
        "crs": _crs(),
    }
    r = requests.get(url, params=params)
    if r.status_code != 200:
        logger.error("[%s] by-TOID error body: %s", collection, r.text)
    r.raise_for_status()
    features = r.json().get("features", [])

    return shape(features[0]["geometry"]) if features else None


def get_nearby_polygons(collection, x, y, pad=SEARCH_PAD_M):
    bbox = f"{x - pad},{y - pad},{x + pad},{y + pad}"
    url = f"{NGD_BASE}/{collection}/items"
    params = {
        "key": OS_API_KEY,
        "bbox": bbox,
        "bbox-crs": _crs(),
        "crs": _crs(),
        "limit": 100,
    }
    r = requests.get(url, params=params)
    if r.status_code != 200:
        logger.error("[%s] bbox error body: %s", collection, r.text)
    r.raise_for_status()

    return r.json().get("features", [])


def find_containing_polygon(x, y, features, fallback_max_m=LAND_CONTAINMENT_FALLBACK_M):
    """Return (toid, polygon) of whichever feature's polygon contains point (x, y).
    Falls back to the nearest polygon within fallback_max_m if no exact
    containment is found (handles address points sitting on/near an edge)."""
    pt = Point(x, y)
    best_toid, best_poly, best_dist = None, None, None
    for feat in features:
        poly = shape(feat["geometry"])
        if poly.covers(pt):
            return feat["properties"]["toid"], poly
        d = poly.distance(pt)
        if best_dist is None or d < best_dist:
            best_toid, best_poly, best_dist = feat["properties"]["toid"], poly, d

    if best_dist is not None and best_dist < fallback_max_m:
        logger.info("No exact containing parcel found, using nearest (%.2fm away)", best_dist)
        return best_toid, best_poly
    return None, None


# ---------------------------------------------------------------------------
# STEP 3 — Real geometric adjacency test
# ---------------------------------------------------------------------------
def shares_real_boundary(poly_a, poly_b, min_shared_length=MIN_SHARED_BOUNDARY_LENGTH_M):
    """True only if the two polygons share an actual boundary LINE (not
    just a single touching corner point), of at least min_shared_length.
    Returns the shared geometry itself (for further checks) or None."""
    intersection = poly_a.boundary.intersection(poly_b.boundary)

    if intersection.is_empty:
        return None
    if intersection.geom_type in ("Point", "MultiPoint"):
        return None

    if hasattr(intersection, "geoms"):
        total_len = sum(g.length for g in intersection.geoms if hasattr(g, "length"))
    else:
        total_len = intersection.length

    if total_len < min_shared_length:
        return None

    return intersection


def is_clean_two_party_boundary(shared_segment, target_toid, candidate_toid,
                                  all_land_features,
                                  min_distance_from_third=PINCH_POINT_CHECK_DISTANCE_M,
                                  min_third_party_area=MIN_THIRD_PARTY_AREA_SQM):
    """Reject a shared boundary if a THIRD property's parcel sits right at
    its MIDPOINT — a signature of a multi-way pinch point (3+ plots meeting
    at a corner) rather than a genuine two-property boundary. Small parcels
    (driveway/path slivers) are ignored so they don't wrongly veto a real
    neighbour relationship."""
    midpoint = shared_segment.interpolate(0.5, normalized=True)
    for feat in all_land_features:
        toid = feat["properties"]["toid"]
        if toid in (target_toid, candidate_toid):
            continue
        poly = shape(feat["geometry"])
        if poly.area < min_third_party_area:
            continue
        if midpoint.distance(poly) < min_distance_from_third:
            return False

    return True


def find_touching_polygons(target_toid, target_polygon, nearby_features, all_land_features=None):
    buffered_target = target_polygon.buffer(BUFFER_M)
    touching = []
    for feat in nearby_features:
        toid = feat["properties"]["toid"]
        if toid == target_toid:
            continue
        poly = shape(feat["geometry"])

        if not buffered_target.intersects(poly):
            continue

        shared = shares_real_boundary(target_polygon, poly)
        if shared is None:
            continue

        if all_land_features is not None:
            if not is_clean_two_party_boundary(shared, target_toid, toid, all_land_features):
                continue

        touching.append({"toid": toid, "polygon": poly})

    return touching


# ---------------------------------------------------------------------------
# STEP 4 — Nearby address points (for matching polygons back to addresses)
# ---------------------------------------------------------------------------
def get_nearby_address_points(x, y, radius=SEARCH_PAD_M):
    url = f"{PLACES_BASE}/radius"
    params = {"point": f"{x},{y}", "radius": radius, "key": OS_API_KEY}
    r = requests.get(url, params=params)
    r.raise_for_status()

    return [res["DPA"] for res in r.json().get("results", []) if "DPA" in res]


def match_address_to_polygon(polygon, address_points,
                              buffer_m=ADDRESS_POINT_SLACK_M,
                              fallback_max_m=ADDRESS_MATCH_FALLBACK_M,
                              context=""):
    """Find address point(s) belonging to this polygon. Tries a buffered
    .covers() test first, then falls back to the single nearest address
    point if it's genuinely close (handles edge-of-polygon placements).

    `context` is a short label (e.g. the candidate's toid) used only for
    logging, so a failure to match is diagnosable rather than a silent
    dead end — logs the nearest distance actually found, or that there
    were no address points in range at all.
    """
    label = f" [{context}]" if context else ""

    poly_buffered = polygon.buffer(buffer_m)
    matched = [dpa for dpa in address_points
               if poly_buffered.covers(Point(dpa["X_COORDINATE"], dpa["Y_COORDINATE"]))]
    if matched:
        return matched

    if not address_points:
        logger.info("No address points at all within search radius%s", label)
        return []

    best_dpa, best_dist = None, None
    for dpa in address_points:
        pt = Point(dpa["X_COORDINATE"], dpa["Y_COORDINATE"])
        d = polygon.distance(pt)
        if best_dist is None or d < best_dist:
            best_dpa, best_dist = dpa, d

    if best_dist is not None and best_dist < fallback_max_m:
        return [best_dpa]

    if best_dist is not None:
        logger.info(
            "No address point matched polygon%s — nearest was %.1fm away "
            "(covers-buffer %.1fm, fallback max %.1fm)",
            label, best_dist, buffer_m, fallback_max_m,
        )

    return []


def sanity_check_distance(target_x, target_y, candidate_dpa,
                           max_dist=MAX_PLAUSIBLE_NEIGHBOUR_DISTANCE_M):
    """Final straight-line check between the target's and candidate's actual
    address points. Catches cases the polygon logic can't reliably detect —
    e.g. two properties both bordering a large shared/communal parcel
    (a cul-de-sac turning circle, a verge) that makes them register as
    'touching' even though they are nowhere near each other in reality."""
    cand_pt = Point(candidate_dpa["X_COORDINATE"], candidate_dpa["Y_COORDINATE"])
    target_pt = Point(target_x, target_y)
    d = target_pt.distance(cand_pt)

    return d <= max_dist, d


# ---------------------------------------------------------------------------
# STEP 5 — Client's filtering rules
# ---------------------------------------------------------------------------
def is_valid_neighbour(dpa_record):
    code = dpa_record.get("CLASSIFICATION_CODE", "")
    state = dpa_record.get("BLPU_STATE_CODE_DESCRIPTION", "")

    if not code.startswith("R"):
        return False, "not residential"
    if code in EXCLUDED_RESIDENTIAL_SUBTYPES:
        return False, "flat/maisonette excluded"
    if state and state.lower() != "in use":
        return False, "not an active/in-use property"

    return True, "ok"


# ---------------------------------------------------------------------------
# MAIN PIPELINE
# ---------------------------------------------------------------------------
def find_neighbours(address_text, debug=False, reference_id=None):
    """`reference_id` ties every row this address produces (across the
    neighbours / full-results / unmatched CSVs and xlsx exports) back to
    this single lookup. Pass one in for batch runs (see main loop below,
    which uses "<run_stamp>-<row_number>"); if omitted, a random one is
    generated so standalone calls still get a usable id."""
    reference_id = reference_id or f"REF-{uuid.uuid4().hex[:10]}"
    logger.info("Processing address [%s]: %s", reference_id, address_text)

    diagnostics = {}
    target = geocode_address(address_text, diagnostics=diagnostics)
    if not target:
        logger.error("Target address not found: %s", address_text)
        return {
            "error": "target address not found",
            "input_address": address_text,
            "diagnostics": diagnostics,
            "reference_id": reference_id,
        }

    target_x, target_y = target["X_COORDINATE"], target["Y_COORDINATE"]
    target_building_toid = target.get("TOPOGRAPHY_LAYER_TOID")
    target_lat, target_lon = xy_to_latlon(target_x, target_y)

    # ---- Layer 1: buildings ----
    target_building_poly = None
    if target_building_toid:
        target_building_poly = get_polygon_by_toid("bld-fts-buildingpart-1", target_building_toid)

    nearby_buildings = get_nearby_polygons("bld-fts-buildingpart-1", target_x, target_y)

    touching_building = []
    if target_building_poly is not None:
        touching_building = find_touching_polygons(
            target_building_toid, target_building_poly, nearby_buildings
        )

    # ---- Layer 2: land parcels (pinch-point veto applied) ----
    nearby_land = get_nearby_polygons("lnd-fts-land-1", target_x, target_y)
    target_land_toid, target_land_poly = find_containing_polygon(target_x, target_y, nearby_land)

    touching_land = []
    if target_land_poly is not None:
        touching_land = find_touching_polygons(
            target_land_toid, target_land_poly, nearby_land,
            all_land_features=nearby_land
        )

    if debug:
        logger.debug("target_building_toid: %s", target_building_toid)
        logger.debug("target_building_poly is None? %s", target_building_poly is None)
        logger.debug("nearby_buildings count: %d", len(nearby_buildings))
        logger.debug("nearby_land count: %d", len(nearby_land))
        logger.debug("target_land_toid: %s", target_land_toid)
        logger.debug("touching_building count: %d", len(touching_building))
        logger.debug("touching_land count: %d", len(touching_land))

    all_touching = touching_building + touching_land

    address_points = get_nearby_address_points(target_x, target_y)

    target_thoroughfare = (target.get("THOROUGHFARE_NAME") or "").strip().upper()

    neighbours = []
    rejected = []
    seen_uprns = set()

    for t in all_touching:
        matched = match_address_to_polygon(
            t["polygon"], address_points, context=f"candidate toid {t['toid']}"
        )

        if not matched:
            rejected.append({"toid": t["toid"], "reason": "no address point found inside polygon"})
            continue

        for dpa in matched:
            uprn = dpa["UPRN"]
            if uprn == target.get("UPRN") or uprn in seen_uprns:
                continue

            # Final sanity check — catches shared-communal-parcel false positives
            ok, dist = sanity_check_distance(target_x, target_y, dpa)
            if not ok:
                seen_uprns.add(uprn)
                rejected.append({
                    "toid": t["toid"], "uprn": uprn, "address": dpa.get("ADDRESS"),
                    "reason": f"address point {dist:.1f}m away — exceeds plausible neighbour distance"
                })
                continue

            # Same-road filter — catches a target's REAR garden parcel touching a
            # property on a completely different street (back-to-back gardens),
            # which the pinch-point/distance checks don't reliably catch since
            # the touching parcel is real and often not that far away.
            candidate_thoroughfare = (dpa.get("THOROUGHFARE_NAME") or "").strip().upper()
            if target_thoroughfare and candidate_thoroughfare and candidate_thoroughfare != target_thoroughfare:
                seen_uprns.add(uprn)
                rejected.append({
                    "toid": t["toid"], "uprn": uprn, "address": dpa.get("ADDRESS"),
                    "reason": (
                        f"different road ('{dpa.get('THOROUGHFARE_NAME')}' vs target's "
                        f"'{target.get('THOROUGHFARE_NAME')}') — likely rear/side parcel "
                        f"touch, not a true neighbour"
                    ),
                })
                continue

            seen_uprns.add(uprn)

            valid, reason = is_valid_neighbour(dpa)
            if valid:
                n_lat, n_lon = xy_to_latlon(dpa["X_COORDINATE"], dpa["Y_COORDINATE"])
                neighbours.append({
                    "uprn": uprn,
                    "address": dpa["ADDRESS"],
                    "classification": dpa.get("CLASSIFICATION_CODE_DESCRIPTION"),
                    "toid": t["toid"],
                    "latitude": n_lat,
                    "longitude": n_lon,
                    "distance_m": round(dist, 1),
                })
            else:
                rejected.append({
                    "toid": t["toid"], "uprn": uprn,
                    "address": dpa.get("ADDRESS"), "reason": reason,
                })

    target_is_flat = target.get("CLASSIFICATION_CODE") in EXCLUDED_RESIDENTIAL_SUBTYPES
    aborted = target_is_flat and len(neighbours) > MAX_NEIGHBOURS_BEFORE_ABORT

    logger.info(
        "Finished %s -> %d neighbour(s), %d rejected candidate(s)%s",
        address_text, len(neighbours), len(rejected),
        " [ABORTED: multi-dwelling]" if aborted else "",
    )

    return {
        "reference_id": reference_id,
        "target": {
            "uprn": target["UPRN"],
            "address": target["ADDRESS"],
            "building_toid": target_building_toid,
            "land_toid": target_land_toid,
            "classification": target.get("CLASSIFICATION_CODE_DESCRIPTION"),
            "latitude": target_lat,
            "longitude": target_lon,
        },
        "neighbours": neighbours,
        "rejected_candidates": rejected,
        "aborted_multi_dwelling": aborted,
        "neighbour_count": len(neighbours),
    }


# ---------------------------------------------------------------------------
# DEBUG HELPER — inspect WHY a boundary is being vetoed, before tuning
# MIN_THIRD_PARTY_AREA_SQM or PINCH_POINT_CHECK_DISTANCE_M.
# ---------------------------------------------------------------------------
def debug_clean_boundary_verbose(target_land_toid, candidate_toid, nearby_land_feats, top_n=5):
    target_poly = None
    cand_poly = None
    for feat in nearby_land_feats:
        if feat["properties"]["toid"] == target_land_toid:
            target_poly = shape(feat["geometry"])
        if feat["properties"]["toid"] == candidate_toid:
            cand_poly = shape(feat["geometry"])

    if target_poly is None or cand_poly is None:
        logger.warning("Could not find one or both polygons in nearby_land_feats")
        return

    shared = target_poly.boundary.intersection(cand_poly.boundary)
    if shared.is_empty or shared.geom_type in ("Point", "MultiPoint"):
        logger.info("No real shared boundary between these two.")
        return

    midpoint = shared.interpolate(0.5, normalized=True)
    logger.info("Shared boundary length: %.2fm", shared.length)

    distances = []
    for feat in nearby_land_feats:
        toid = feat["properties"]["toid"]
        if toid in (target_land_toid, candidate_toid):
            continue
        poly = shape(feat["geometry"])
        d = midpoint.distance(poly)
        distances.append((d, toid, poly.area))

    distances.sort(key=lambda x: x[0])
    logger.info("Nearest %d other parcels to this boundary's midpoint:", top_n)
    for d, toid, area in distances[:top_n]:
        logger.info("  %s: %.3fm away, area=%.1f sqm", toid, d, area)


# ---------------------------------------------------------------------------
# CSV WRITERS — incremental (append one address's results at a time) so a
# crash mid-batch doesn't lose everything already processed.
# ---------------------------------------------------------------------------
def _init_csv(path, fieldnames):
    """Create the CSV with a header row if it doesn't already exist."""
    if not os.path.exists(path):
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
        logger.debug("Initialised CSV: %s", path)


def append_neighbours_csv(result, path=NEIGHBOURS_CSV_PATH):
    """Append one row per confirmed neighbour. Writes nothing if there's
    no target (geocode failure) or no neighbours."""
    if "error" in result or not result.get("neighbours"):
        return

    _init_csv(path, NEIGHBOURS_CSV_FIELDS)
    target = result["target"]

    with open(path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=NEIGHBOURS_CSV_FIELDS)
        for n in result["neighbours"]:
            writer.writerow({
                "Reference Id": result.get("reference_id", ""),
                "Target Uprn": target["uprn"],
                "Target Address": target["address"],
                "Target Building Toid": target["building_toid"],
                "Target Land Toid": target["land_toid"],
                "Target Classification": target["classification"],
                "Target Latitude": target["latitude"],
                "Target Longitude": target["longitude"],
                "Neighbour Uprn": n["uprn"],
                "Neighbour Address": n["address"],
                "Neighbour Classification": n["classification"],
                "Neighbour Toid": n["toid"],
                "Neighbour Latitude": n["latitude"],
                "Neighbour Longitude": n["longitude"],
                "Distance M": n["distance_m"],
            })

    logger.debug("Appended %d row(s) to %s", len(result["neighbours"]), path)


def append_full_results_csv(result, input_address=None, path=FULL_RESULTS_CSV_PATH):
    """Append every candidate the pipeline looked at for this address:
    confirmed neighbours, rejected candidates, and (if neither) a single
    'None Found' or 'Error' row so the address is still represented in the
    audit trail even when nothing came out of it."""
    _init_csv(path, FULL_RESULTS_CSV_FIELDS)

    if "error" in result:
        with open(path, "a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=FULL_RESULTS_CSV_FIELDS)
            writer.writerow({
                "Reference Id": result.get("reference_id", ""),
                "Target Uprn": "", "Target Address": input_address or "",
                "Target Building Toid": "", "Target Land Toid": "",
                "Target Classification": "", "Target Latitude": "", "Target Longitude": "",
                "Aborted Multi Dwelling": "", "Neighbour Count": "", "Record Type": "Error",
                "Uprn": "", "Address": "", "Classification": "", "Toid": "",
                "Latitude": "", "Longitude": "", "Distance M": "",
                "Reason": result.get("error", ""),
            })
        logger.debug("Appended error row to %s for %s", path, input_address)
        return

    target = result["target"]
    base_row = {
        "Reference Id": result.get("reference_id", ""),
        "Target Uprn": target["uprn"],
        "Target Address": target["address"],
        "Target Building Toid": target["building_toid"],
        "Target Land Toid": target["land_toid"],
        "Target Classification": target["classification"],
        "Target Latitude": target["latitude"],
        "Target Longitude": target["longitude"],
        "Aborted Multi Dwelling": result["aborted_multi_dwelling"],
        "Neighbour Count": result["neighbour_count"],
    }

    rows = []

    for n in result["neighbours"]:
        rows.append({
            **base_row,
            "Record Type": "Neighbour",
            "Uprn": n["uprn"], "Address": n["address"],
            "Classification": n["classification"], "Toid": n["toid"],
            "Latitude": n["latitude"], "Longitude": n["longitude"],
            "Distance M": n["distance_m"], "Reason": "",
        })

    for rej in result["rejected_candidates"]:
        rows.append({
            **base_row,
            "Record Type": "Rejected",
            "Uprn": rej.get("uprn", ""), "Address": rej.get("address", ""),
            "Classification": "", "Toid": rej.get("toid", ""),
            "Latitude": "", "Longitude": "",
            "Distance M": "", "Reason": rej.get("reason", ""),
        })

    if not rows:
        rows.append({
            **base_row,
            "Record Type": "None Found",
            "Uprn": "", "Address": "", "Classification": "", "Toid": "",
            "Latitude": "", "Longitude": "",
            "Distance M": "", "Reason": "no candidate polygons touched the target",
        })

    with open(path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FULL_RESULTS_CSV_FIELDS)
        writer.writerows(rows)

    logger.debug("Appended %d row(s) to %s", len(rows), path)


def append_unmatched_csv(result, path=UNMATCHED_CSV_PATH):
    """Append one row for an address that failed to geocode, with whatever
    partial match info OS Places returned, so failed inputs are triageable
    instead of showing up as a blank row."""
    if "error" not in result:
        return

    _init_csv(path, UNMATCHED_CSV_FIELDS)
    diag = result.get("diagnostics", {})

    with open(path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=UNMATCHED_CSV_FIELDS)
        writer.writerow({
            "Reference Id": result.get("reference_id", ""),
            "Input Address": result.get("input_address", ""),
            "Raw Results Count": diag.get("raw_results_count", ""),
            "Best Match Address": diag.get("best_match_address", ""),
            "Best Match Score": diag.get("best_match_score", ""),
        })

    logger.debug("Appended unmatched-address row to %s for %s", path, result.get("input_address"))


# ---------------------------------------------------------------------------
# XLSX EXPORT — Title Case, bold header row. The CSVs above are the
# incremental audit trail written during the run; this builds the polished
# .xlsx deliverable from a finished CSV once the batch completes.
# ---------------------------------------------------------------------------
def write_formatted_xlsx(csv_path, xlsx_path, sheet_name="Sheet1"):
    """Load a CSV (already using our Title Case headers) and write it out
    as a formatted .xlsx: bold header row, frozen header, auto-sized
    columns. No-ops if the CSV doesn't exist (e.g. no neighbours found)."""
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    if not os.path.exists(csv_path):
        logger.info("Skipping xlsx export — %s does not exist", csv_path)
        return None

    df = pd.read_csv(csv_path)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for row in df.itertuples(index=False):
        ws.append(list(row))

    # Auto-size columns based on the longest value (header or data) in each.
    for col_idx, col_name in enumerate(df.columns, start=1):
        longest = max(
            [len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str)]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(longest + 2, 60)

    wb.save(xlsx_path)
    logger.info("Wrote formatted xlsx: %s", xlsx_path)
    return xlsx_path


if __name__ == "__main__":
    import pandas as pd

    logger.info("Starting run. Log file: %s", LOG_FILE_PATH)
    logger.info("Neighbours CSV: %s", NEIGHBOURS_CSV_PATH)
    logger.info("Full results CSV: %s", FULL_RESULTS_CSV_PATH)
    logger.info("Unmatched addresses CSV: %s", UNMATCHED_CSV_PATH)

    data = pd.read_csv(INPUT_FILE_PATH)
    logger.info("Loaded %d address(es) from %s", len(data), INPUT_FILE_PATH)

    # If the input file already carries its own reference/id column, use
    # that (so output rows tie back to the client's own record); otherwise
    # generate one per row: "<run_stamp>-<row_number>".

    processed, failed, unmatched = 0, 0, 0

    for index, row in data.iterrows():
        address = row["Address"]
        reference_id = row["Reference"]
        logger.info(f"{index} is on working.....")
        try:
            result = find_neighbours(address, debug=True, reference_id=reference_id)
            append_full_results_csv(result, input_address=address)
            if "error" in result:
                append_unmatched_csv(result)
                unmatched += 1
            else:
                append_neighbours_csv(result)
            processed += 1
        except Exception:
            logger.exception("Unhandled error processing address: %s", address)
            error_result = {
                "error": "unhandled exception",
                "input_address": address,
                "reference_id": reference_id,
            }
            append_full_results_csv(error_result, input_address=address)
            append_unmatched_csv(error_result)
            failed += 1

    logger.info(
        "Run complete. %d processed, %d failed, %d unmatched (geocode) addresses.",
        processed, failed, unmatched,
    )
    logger.info("Neighbours CSV written to: %s", NEIGHBOURS_CSV_PATH)
    logger.info("Full results CSV written to: %s", FULL_RESULTS_CSV_PATH)
    logger.info("Unmatched addresses CSV written to: %s", UNMATCHED_CSV_PATH)

    neighbours_xlsx_path = os.path.join(OUTPUT_DIR, f"neighbours_{RUN_STAMP}.xlsx")
    full_results_xlsx_path = os.path.join(OUTPUT_DIR, f"full_results_{RUN_STAMP}.xlsx")
    write_formatted_xlsx(NEIGHBOURS_CSV_PATH, neighbours_xlsx_path, sheet_name="Neighbours")
    write_formatted_xlsx(FULL_RESULTS_CSV_PATH, full_results_xlsx_path, sheet_name="Full Results")